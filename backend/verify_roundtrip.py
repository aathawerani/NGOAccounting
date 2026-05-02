# Round-trip verification: export HTTT full workbook and check sheet/row counts.
import sys
import os
import requests
import openpyxl
import sqlite3

API = "http://localhost:8000"
DB_PATH = os.path.join(os.path.dirname(__file__), "ngo_accounting.db")
OUT_FILE = "HTTT-roundtrip.xlsx"

# ── 1. Get trust_id for HTTT ────────────────────────────────────────────────
resp = requests.get(f"{API}/api/trusts")
resp.raise_for_status()
trusts = resp.json()
httt = next((t for t in trusts if t["code"] == "HTTT"), None)
if not httt:
    print("ERROR: HTTT trust not found via API")
    sys.exit(1)
trust_id = httt["id"]
print(f"HTTT trust_id = {trust_id}")

# ── 2. Export full workbook (mode=full, no year filter = all time) ──────────
print(f"Downloading full workbook from {API}/api/export/full ...")
r = requests.get(f"{API}/api/export/full", params={"trust_id": trust_id, "mode": "full"})
if not r.ok:
    try:
        detail = r.json()
    except Exception:
        detail = r.text
    print(f"ERROR: Export failed ({r.status_code}): {detail}")
    sys.exit(1)
with open(OUT_FILE, "wb") as f:
    f.write(r.content)
print(f"Saved to {OUT_FILE} ({len(r.content):,} bytes)")

# ── 3. Open workbook and inspect ────────────────────────────────────────────
wb = openpyxl.load_workbook(OUT_FILE)
sheet_names = wb.sheetnames
print(f"\nTotal sheets: {len(sheet_names)}")
print(f"Sheets: {sheet_names}")

summary_sheets = [s for s in sheet_names if s in ("TB", "IS", "BS")]
account_sheets = [s for s in sheet_names if s not in ("TB", "IS", "BS")]
print(f"\nAccount sheets: {len(account_sheets)}")
print(f"Summary sheets: {len(summary_sheets)} -> {summary_sheets}")

# ── 4. Count transaction rows across all account sheets ─────────────────────
HEADER_ROW = 13  # transactions start at row 13
total_tx_rows = 0
for sname in account_sheets:
    ws = wb[sname]
    max_row = ws.max_row
    # count non-empty rows from row 13 onwards
    count = 0
    for row_idx in range(HEADER_ROW, max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1).value
        if cell_a is not None and str(cell_a).strip() != "":
            count += 1
    total_tx_rows += count

print(f"\nTransaction rows across all account sheets: {total_tx_rows}")

# ── 5. Compare with DB ───────────────────────────────────────────────────────
con = sqlite3.connect(DB_PATH)
cur = con.cursor()
cur.execute("SELECT COUNT(*) FROM ledger_entries WHERE trust_id = ?", (trust_id,))
db_count = cur.fetchone()[0]
con.close()
print(f"DB ledger_entries for HTTT: {db_count}")

if total_tx_rows == db_count:
    print(f"\nMATCH: {total_tx_rows} rows in export == {db_count} in DB")
else:
    print(f"\nMISMATCH: export has {total_tx_rows}, DB has {db_count}")

# ── 6. Spot-check first account sheet ───────────────────────────────────────
if account_sheets:
    ws = wb[account_sheets[0]]
    print(f"\nSpot-check sheet '{account_sheets[0]}':")
    print(f"  A3  (title):      {ws.cell(3, 1).value!r}")
    print(f"  B5  (label):      {ws.cell(5, 2).value!r}")
    print(f"  A10 (col header): {ws.cell(10, 1).value!r}")
    print(f"  A13 (first tx):   {ws.cell(13, 1).value!r}")
    print(f"  B13 (receipt no): {ws.cell(13, 2).value!r}")
    print(f"  D13 (tenant):     {ws.cell(13, 4).value!r}")

print("\nDone.")
