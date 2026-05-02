"""Excel export: Trial Balance, Ledger, Income Statement, Balance Sheet."""
from io import BytesIO
from datetime import date as date_type
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from database import get_db
from models.models import AccountType, LedgerEntry, Trust

router = APIRouter(prefix="/api/export", tags=["export"])


# ── Style constants ───────────────────────────────────────────────────────────

_DARK_FILL  = PatternFill("solid", fgColor="1E293B")   # slate-800
_BLUE_FILL  = PatternFill("solid", fgColor="0F4C81")   # header row
_ALT_FILL   = PatternFill("solid", fgColor="F1F5F9")   # alternate row
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
_BOLD_WHITE = Font(bold=True, color="FFFFFF", size=11)
_BOLD_DARK  = Font(bold=True, color="1E293B", size=10)
_BOLD_TITLE = Font(bold=True, color="FFFFFF", size=13)
_NORMAL     = Font(size=10)

_THIN = Side(style="thin", color="CBD5E1")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT  = Alignment(horizontal="right", vertical="center")
_LEFT   = Alignment(horizontal="left", vertical="center")

_PKR_FMT = '#,##0'


def _style(cell, fill=None, font=None, align=None, border=True, num_format=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = _BORDER
    if num_format:
        cell.number_format = num_format


def _header_row(ws, row: int, labels: list, fills=None):
    for col, label in enumerate(labels, 1):
        c = ws.cell(row=row, column=col, value=label)
        _style(c, fill=_BLUE_FILL, font=_BOLD_WHITE, align=_CENTER)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)


def _title_block(ws, trust_name: str, date_from: str, date_to: str, title: str, ncols: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=trust_name)
    _style(c, fill=_DARK_FILL, font=_BOLD_TITLE, align=_CENTER)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=title)
    _style(c, fill=_DARK_FILL, font=_BOLD_WHITE, align=_CENTER)

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    period = f"Period: {date_from}  to  {date_to}"
    c = ws.cell(row=3, column=1, value=period)
    _style(c, fill=_DARK_FILL, font=Font(color="94A3B8", italic=True, size=9), align=_CENTER)

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_trial_balance(ws, trust, accounts, entries, date_from, date_to):
    _title_block(ws, trust.name, date_from, date_to, "Trial Balance", 5)
    _header_row(ws, 5, ["Account Code", "Account Name", "Type", "Total Debit", "Total Credit"])

    # Aggregate per account
    from collections import defaultdict
    totals = defaultdict(lambda: {"debit": 0.0, "credit": 0.0})
    for e in entries:
        totals[e.account_code]["debit"] += e.debit
        totals[e.account_code]["credit"] += e.credit

    grand_dr = grand_cr = 0.0
    for row_idx, acct in enumerate(accounts, start=6):
        fill = _ALT_FILL if row_idx % 2 == 0 else _WHITE_FILL
        dr = totals[acct.account_code]["debit"]
        cr = totals[acct.account_code]["credit"]
        grand_dr += dr
        grand_cr += cr
        data = [acct.account_code, acct.account_name, acct.account_type, dr, cr]
        for col, val in enumerate(data, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            num_fmt = _PKR_FMT if col >= 4 else None
            align = _RIGHT if col >= 4 else _LEFT
            _style(c, fill=fill, font=_NORMAL, align=align, num_format=num_fmt)

    total_row = len(accounts) + 6
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=3, value="")
    dr_c = ws.cell(row=total_row, column=4, value=grand_dr)
    cr_c = ws.cell(row=total_row, column=5, value=grand_cr)
    for c in [ws.cell(row=total_row, column=i) for i in range(1, 6)]:
        _style(c, fill=_DARK_FILL, font=_BOLD_WHITE, align=_CENTER)
    dr_c.alignment = _RIGHT
    cr_c.alignment = _RIGHT
    dr_c.number_format = _PKR_FMT
    cr_c.number_format = _PKR_FMT

    _auto_width(ws)


def _build_ledger(ws, trust, accounts, entries, date_from, date_to):
    _title_block(ws, trust.name, date_from, date_to, "General Ledger", 8)

    acct_map = {a.account_code: a for a in accounts}
    from collections import defaultdict
    by_acct = defaultdict(list)
    for e in entries:
        by_acct[e.account_code].append(e)

    row_idx = 5
    _header_row(ws, row_idx, ["Date", "Ref No.", "Party Name", "Contra Account", "Particulars", "Debit", "Credit", "Balance"])
    row_idx += 1

    for acct_code, acct_entries in sorted(by_acct.items()):
        # Section header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)
        label = f"{acct_code} — {acct_map.get(acct_code, acct_code if isinstance(acct_code, str) else acct_code).account_name if acct_code in acct_map else acct_code}"
        c = ws.cell(row=row_idx, column=1, value=label)
        _style(c, fill=_BLUE_FILL, font=_BOLD_WHITE, align=_LEFT)
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

        balance = 0.0
        for i, e in enumerate(sorted(acct_entries, key=lambda x: (x.date, x.id))):
            balance += e.debit - e.credit
            fill = _ALT_FILL if i % 2 == 0 else _WHITE_FILL
            row_data = [
                e.date.isoformat() if e.date else "",
                e.receipt_no or "",
                e.party_name or "",
                e.contra_account_code or "",
                e.particulars or "",
                e.debit,
                e.credit,
                abs(balance),
            ]
            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                num_fmt = _PKR_FMT if col >= 6 else None
                align = _RIGHT if col >= 6 else _LEFT
                _style(c, fill=fill, font=_NORMAL, align=align, num_format=num_fmt)
            row_idx += 1

    _auto_width(ws)


def _build_income_statement(ws, trust, accounts, entries, date_from, date_to):
    _title_block(ws, trust.name, date_from, date_to, "Income Statement", 3)
    row_idx = 5

    from collections import defaultdict
    totals = defaultdict(lambda: {"debit": 0.0, "credit": 0.0})
    for e in entries:
        totals[e.account_code]["debit"] += e.debit
        totals[e.account_code]["credit"] += e.credit

    acct_map = {a.account_code: a for a in accounts}

    income_accts   = [a for a in accounts if a.account_type == "INCOME"]
    expense_accts  = [a for a in accounts if a.account_type == "EXPENSE"]

    def _section(label, accts, flip=False):
        nonlocal row_idx
        # Section header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
        c = ws.cell(row=row_idx, column=1, value=label)
        _style(c, fill=_BLUE_FILL, font=_BOLD_WHITE, align=_LEFT)
        row_idx += 1

        section_total = 0.0
        for i, acct in enumerate(accts):
            dr = totals[acct.account_code]["debit"]
            cr = totals[acct.account_code]["credit"]
            net = cr - dr if not flip else dr - cr
            fill = _ALT_FILL if i % 2 == 0 else _WHITE_FILL
            ws.cell(row=row_idx, column=1, value=acct.account_code).fill = fill
            ws.cell(row=row_idx, column=2, value=acct.account_name).fill = fill
            c_net = ws.cell(row=row_idx, column=3, value=net)
            for col in range(1, 4):
                cc = ws.cell(row=row_idx, column=col)
                _style(cc, fill=fill, font=_NORMAL, align=_RIGHT if col == 3 else _LEFT, num_format=_PKR_FMT if col == 3 else None)
            section_total += net
            row_idx += 1

        # Subtotal
        c = ws.cell(row=row_idx, column=2, value=f"Total {label}")
        t = ws.cell(row=row_idx, column=3, value=section_total)
        for col in range(1, 4):
            cc = ws.cell(row=row_idx, column=col)
            _style(cc, fill=_DARK_FILL, font=_BOLD_WHITE, align=_RIGHT if col >= 2 else _LEFT, num_format=_PKR_FMT if col == 3 else None)
        row_idx += 2
        return section_total

    total_income  = _section("INCOME", income_accts, flip=False)
    total_expense = _section("EXPENSES", expense_accts, flip=True)

    net = total_income - total_expense
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
    c1 = ws.cell(row=row_idx, column=1, value="NET PROFIT / (LOSS)")
    c2 = ws.cell(row=row_idx, column=3, value=net)
    for col in range(1, 4):
        cc = ws.cell(row=row_idx, column=col)
        _style(cc, fill=_DARK_FILL, font=_BOLD_WHITE, align=_RIGHT if col == 3 else _LEFT, num_format=_PKR_FMT if col == 3 else None)

    _auto_width(ws)


def _build_balance_sheet(ws, trust, accounts, entries, date_from, date_to):
    _title_block(ws, trust.name, date_from, date_to, "Balance Sheet", 3)
    row_idx = 5

    from collections import defaultdict
    totals = defaultdict(lambda: {"debit": 0.0, "credit": 0.0})
    for e in entries:
        totals[e.account_code]["debit"] += e.debit
        totals[e.account_code]["credit"] += e.credit

    asset_accts  = [a for a in accounts if a.account_type == "ASSET"]
    liab_accts   = [a for a in accounts if a.account_type == "LIABILITY"]
    equity_accts = [a for a in accounts if a.account_type == "EQUITY"]
    # Net profit from IS goes into equity
    income_accts   = [a for a in accounts if a.account_type == "INCOME"]
    expense_accts  = [a for a in accounts if a.account_type == "EXPENSE"]

    def _section(label, accts, positive_is_cr=False):
        nonlocal row_idx
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
        c = ws.cell(row=row_idx, column=1, value=label)
        _style(c, fill=_BLUE_FILL, font=_BOLD_WHITE, align=_LEFT)
        row_idx += 1
        total = 0.0
        for i, acct in enumerate(accts):
            dr = totals[acct.account_code]["debit"]
            cr = totals[acct.account_code]["credit"]
            net = (cr - dr) if positive_is_cr else (dr - cr)
            fill = _ALT_FILL if i % 2 == 0 else _WHITE_FILL
            for col, val in enumerate([acct.account_code, acct.account_name, net], 1):
                cc = ws.cell(row=row_idx, column=col, value=val)
                _style(cc, fill=fill, font=_NORMAL, align=_RIGHT if col == 3 else _LEFT, num_format=_PKR_FMT if col == 3 else None)
            total += net
            row_idx += 1
        for col in range(1, 4):
            cc = ws.cell(row=row_idx, column=col)
            cc.value = f"Total {label}" if col == 2 else (total if col == 3 else "")
            _style(cc, fill=_DARK_FILL, font=_BOLD_WHITE, align=_RIGHT if col >= 2 else _LEFT, num_format=_PKR_FMT if col == 3 else None)
        row_idx += 2
        return total

    total_assets  = _section("ASSETS", asset_accts, positive_is_cr=False)
    total_liab    = _section("LIABILITIES", liab_accts, positive_is_cr=True)
    total_equity  = _section("EQUITY / FUND", equity_accts, positive_is_cr=True)

    # Net profit
    income_tot  = sum((totals[a.account_code]["credit"] - totals[a.account_code]["debit"]) for a in income_accts)
    expense_tot = sum((totals[a.account_code]["debit"] - totals[a.account_code]["credit"]) for a in expense_accts)
    net_profit  = income_tot - expense_tot

    for col, val in enumerate(["", "Net Profit / (Loss)", net_profit], 1):
        cc = ws.cell(row=row_idx, column=col, value=val)
        _style(cc, fill=_ALT_FILL, font=_BOLD_DARK, align=_RIGHT if col == 3 else _LEFT, num_format=_PKR_FMT if col == 3 else None)
    row_idx += 2

    total_l_e = total_liab + total_equity + net_profit
    for col, val in enumerate(["", "TOTAL LIABILITIES & EQUITY", total_l_e], 1):
        cc = ws.cell(row=row_idx, column=col, value=val)
        _style(cc, fill=_DARK_FILL, font=_BOLD_WHITE, align=_RIGHT if col == 3 else _LEFT, num_format=_PKR_FMT if col == 3 else None)

    _auto_width(ws)


# ── Main export endpoint ──────────────────────────────────────────────────────

@router.get("/reports")
def export_reports(
    trust_id: int,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
):
    trust = db.query(Trust).filter(Trust.id == trust_id).first()
    if not trust:
        raise HTTPException(404, "Trust not found")

    d_from = date_from or "2000-01-01"
    d_to   = date_to   or date_type.today().isoformat()
    label_from = d_from
    label_to   = d_to

    try:
        from datetime import datetime
        d_from_obj = datetime.strptime(d_from, "%Y-%m-%d").date()
        d_to_obj   = datetime.strptime(d_to,   "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, "Dates must be YYYY-MM-DD")

    accounts = (
        db.query(AccountType)
        .filter(AccountType.trust_id == trust_id)
        .order_by(AccountType.account_type, AccountType.account_code)
        .all()
    )
    entries = (
        db.query(LedgerEntry)
        .filter(
            LedgerEntry.trust_id == trust_id,
            LedgerEntry.date >= d_from_obj,
            LedgerEntry.date <= d_to_obj,
        )
        .order_by(LedgerEntry.date, LedgerEntry.id)
        .all()
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    ws_tb = wb.create_sheet("Trial Balance")
    _build_trial_balance(ws_tb, trust, accounts, entries, label_from, label_to)

    ws_gl = wb.create_sheet("General Ledger")
    _build_ledger(ws_gl, trust, accounts, entries, label_from, label_to)

    ws_is = wb.create_sheet("Income Statement")
    _build_income_statement(ws_is, trust, accounts, entries, label_from, label_to)

    ws_bs = wb.create_sheet("Balance Sheet")
    _build_balance_sheet(ws_bs, trust, accounts, entries, label_from, label_to)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{trust.code}_reports_{d_from}_{d_to}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── WPF-format per-account ledger export ─────────────────────────────────────

_WPF_HEADER_FILL  = PatternFill("solid", fgColor="4472C4")   # Excel blue
_WPF_TITLE_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
_WPF_SUBHDR_FILL  = PatternFill("solid", fgColor="D9E1F2")   # light blue
_WPF_LABEL_FONT   = Font(bold=True, size=10)
_WPF_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
_WPF_TITLE_FONT   = Font(bold=True, color="FFFFFF", size=12)
_WPF_NORMAL_FONT  = Font(size=10)
_WPF_THIN         = Side(style="thin", color="BFBFBF")
_WPF_BORDER       = Border(left=_WPF_THIN, right=_WPF_THIN,
                           top=_WPF_THIN, bottom=_WPF_THIN)


def _wpf_account_sheet(ws, account: AccountType, entries: list, trust_name: str):
    """
    Write a single account sheet in WPF format.

    Row layout (openpyxl 1-indexed):
      1-2  : empty
      3    : "GENERAL LEDGER" merged across cols A-H
      4    : empty
      5    : col B = "NAME OF ACCOUNT :"   col E = account_name
      6    : col B = "TYPE OF ACCOUNT:"    col E = account_type
      7    : col B = "ACCOUNT CODE  :"     col E = account_code
      8-9  : empty
      10   : column headers
      11-12: empty
      13+  : transaction rows
    """
    # Row 3 — title
    ws.merge_cells("A3:H3")
    c = ws["A3"]
    c.value = "GENERAL LEDGER"
    c.font = _WPF_TITLE_FONT
    c.fill = _WPF_TITLE_FILL
    c.alignment = _CENTER
    ws.row_dimensions[3].height = 22

    # Rows 5-7 — account metadata
    meta_rows = [
        (5, "NAME OF ACCOUNT :", account.account_name),
        (6, "TYPE OF ACCOUNT:",  account.account_type),
        (7, "ACCOUNT CODE  :",   account.account_code),
    ]
    for r, label, value in meta_rows:
        lc = ws.cell(row=r, column=2, value=label)
        lc.font = _WPF_LABEL_FONT
        lc.fill = _WPF_SUBHDR_FILL
        lc.alignment = _RIGHT
        vc = ws.cell(row=r, column=5, value=value)
        vc.font = _WPF_LABEL_FONT
        vc.fill = _WPF_SUBHDR_FILL

    # Row 10 — column headers
    headers = ["DATE", "RECEIPT/VOUCHER NO", "ACCOUNT CODE",
               "NAME OF TENANT", "PARTICULARS", "DEBIT", "CREDIT", "BALANCE"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=10, column=col, value=h)
        c.font = _WPF_HEADER_FONT
        c.fill = _WPF_HEADER_FILL
        c.alignment = _CENTER
        c.border = _WPF_BORDER
    ws.row_dimensions[10].height = 18

    # Rows 13+ — transactions with running balance
    balance = 0.0
    for i, e in enumerate(sorted(entries, key=lambda x: (x.date, x.id))):
        balance += e.debit - e.credit
        row_num = 13 + i
        fill = _ALT_FILL if i % 2 else _WHITE_FILL
        values = [
            e.date.strftime("%d-%m-%Y") if e.date else "",
            e.receipt_no or "",
            e.contra_account_code or "",
            e.party_name or "",
            e.particulars or "",
            e.debit if e.debit else None,
            e.credit if e.credit else None,
            abs(balance),
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.font = _WPF_NORMAL_FONT
            c.fill = fill
            c.border = _WPF_BORDER
            if col in (6, 7, 8):
                c.number_format = _PKR_FMT
                c.alignment = _RIGHT
            else:
                c.alignment = _LEFT

    # Column widths
    col_widths = [14, 20, 16, 30, 40, 14, 14, 16]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze panes below header
    ws.freeze_panes = "A13"


@router.get("/ledger")
def export_ledger(
    trust_id: int,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """
    Export every account as its own sheet in WPF format, plus TB / IS / BS summary sheets.
    ?year=2024 filters to that fiscal year (Jul–Jun). Omit for all data.
    """
    trust = db.query(Trust).filter(Trust.id == trust_id).first()
    if not trust:
        raise HTTPException(404, "Trust not found")

    from datetime import datetime as _dt
    if year:
        d_from_obj = _dt(year - 1, 7, 1).date()
        d_to_obj   = _dt(year, 6, 30).date()
        label_from = d_from_obj.isoformat()
        label_to   = d_to_obj.isoformat()
    else:
        d_from_obj = _dt(2000, 1, 1).date()
        d_to_obj   = date_type.today()
        label_from = "All time"
        label_to   = d_to_obj.isoformat()

    accounts = (
        db.query(AccountType)
        .filter(AccountType.trust_id == trust_id)
        .order_by(AccountType.account_type, AccountType.account_code)
        .all()
    )
    all_entries = (
        db.query(LedgerEntry)
        .filter(
            LedgerEntry.trust_id == trust_id,
            LedgerEntry.date >= d_from_obj,
            LedgerEntry.date <= d_to_obj,
        )
        .order_by(LedgerEntry.date, LedgerEntry.id)
        .all()
    )

    from collections import defaultdict
    by_acct: dict = defaultdict(list)
    for e in all_entries:
        by_acct[e.account_code].append(e)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # One sheet per account
    for acct in accounts:
        # Sheet names must be ≤31 chars and no invalid chars
        safe_name = acct.account_code[:31].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "").replace("]", "")
        ws = wb.create_sheet(safe_name)
        _wpf_account_sheet(ws, acct, by_acct.get(acct.account_code, []), trust.name)

    # Summary sheets at the end (using the styled report builders)
    ws_tb = wb.create_sheet("TB")
    _build_trial_balance(ws_tb, trust, accounts, all_entries, label_from, label_to)

    ws_is = wb.create_sheet("IS")
    _build_income_statement(ws_is, trust, accounts, all_entries, label_from, label_to)

    ws_bs = wb.create_sheet("BS")
    _build_balance_sheet(ws_bs, trust, accounts, all_entries, label_from, label_to)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    yr_label = str(year) if year else "all"
    filename = f"{trust.code}{yr_label}-ledger.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── WPF exact-format full export ──────────────────────────────────────────────
#
# Colors and column layout reproduce the original WPF Excel file so the output
# can be re-imported by the same import pipeline without any changes.

_F_DARK_FILL   = PatternFill("solid", fgColor="1A5276")   # dark green  #1a5276
_F_SUBHDR_FILL = PatternFill("solid", fgColor="D5D8DC")   # light gray  #d5d8dc
_F_ALT_FILL    = PatternFill("solid", fgColor="EAFAF1")    # light green #eafaf1
_F_WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
_F_WHITE_FONT  = Font(bold=True, color="FFFFFF", size=10)
_F_HDR_FONT    = Font(bold=True, color="1E293B", size=10)
_F_NORMAL_FONT = Font(size=10)
_F_COL_WIDTHS  = [14, 14, 12, 25, 45, 14, 14, 14]   # per task spec


def _f_row(ws, row: int, fill, values: list, num_cols=(6, 7, 8), bold=False):
    """Write a styled data row. num_cols are 1-indexed columns to right-align + number-format."""
    font = Font(bold=bold, size=10) if not bold else _F_HDR_FONT
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = _F_WHITE_FONT if fill is _F_DARK_FILL else (_F_HDR_FONT if fill is _F_SUBHDR_FILL else _F_NORMAL_FONT)
        c.fill = fill
        c.border = _BORDER
        c.alignment = _RIGHT if col in num_cols else _LEFT
        if col in num_cols and val is not None:
            c.number_format = _PKR_FMT


def _wpf_full_account_sheet(ws, account: AccountType, entries: list, _trust_name: str):
    """
    Account sheet in exact WPF format:
      Row 3  : GENERAL LEDGER (merged A3:H3) — dark green
      Rows 5-7: metadata labels in col B, values in col E — dark green
      Row 10 : column headers — light gray
      Row 13+: transactions — alternating white / light green
    """
    # Row 3 — title
    ws.merge_cells("A3:H3")
    c = ws["A3"]
    c.value, c.font, c.fill, c.alignment = "GENERAL LEDGER", _F_WHITE_FONT, _F_DARK_FILL, _CENTER
    ws.row_dimensions[3].height = 22

    # Rows 5-7 — account metadata
    for r, label, value in [
        (5, "NAME OF ACCOUNT :", account.account_name),
        (6, "TYPE OF ACCOUNT:",  account.account_type),
        (7, "ACCOUNT CODE  :",   account.account_code),
    ]:
        for col in range(1, 9):  # colour entire row dark green
            ws.cell(row=r, column=col).fill = _F_DARK_FILL
        lc = ws.cell(row=r, column=2, value=label)
        lc.font, lc.alignment = _F_WHITE_FONT, _RIGHT
        vc = ws.cell(row=r, column=5, value=value)
        vc.font = _F_WHITE_FONT

    # Row 10 — column headers
    headers = ["DATE", "RECEIPT / VOUCHER NO", "ACCOUNT CODE",
               "NAME OF TENANT", "PARTICULARS", "DEBIT", "CREDIT", "BALANCE"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=10, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = _F_HDR_FONT, _F_SUBHDR_FILL, _CENTER, _BORDER
    ws.row_dimensions[10].height = 18

    # Rows 13+ — transactions
    balance = 0.0
    for i, e in enumerate(sorted(entries, key=lambda x: (x.date, x.id))):
        balance += e.debit - e.credit
        fill = _F_WHITE_FILL if i % 2 == 0 else _F_ALT_FILL
        values = [
            e.date.strftime("%d/%m/%Y") if e.date else "",
            e.receipt_no or "",
            e.contra_account_code or "",
            e.party_name or "",
            e.particulars or "",
            e.debit   if e.debit   else None,
            e.credit  if e.credit  else None,
            abs(balance),
        ]
        _f_row(ws, 13 + i, fill, values, num_cols=(6, 7, 8))

    # Column widths + freeze
    for col, w in enumerate(_F_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A13"


def _wpf_full_tb_sheet(ws, trust, accounts, entries, label_from, label_to):
    from collections import defaultdict
    totals = defaultdict(lambda: {"dr": 0.0, "cr": 0.0})
    for e in entries:
        totals[e.account_code]["dr"] += e.debit
        totals[e.account_code]["cr"] += e.credit

    # Title block (3 rows)
    for r, val in [(1, trust.name), (2, "TRIAL BALANCE"), (3, f"Period: {label_from}  –  {label_to}")]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value=val)
        c.font = _F_WHITE_FONT if r < 3 else Font(italic=True, color="FFFFFF", size=9)
        c.fill, c.alignment = _F_DARK_FILL, _CENTER

    # Column headers row 5
    for col, h in enumerate(["Code", "Account Name", "Type", "Debit", "Credit", "Net Balance"], 1):
        c = ws.cell(row=5, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = _F_HDR_FONT, _F_SUBHDR_FILL, _CENTER, _BORDER

    grand_dr = grand_cr = 0.0
    for i, acct in enumerate(accounts, start=6):
        fill = _F_WHITE_FILL if i % 2 == 0 else _F_ALT_FILL
        dr, cr = totals[acct.account_code]["dr"], totals[acct.account_code]["cr"]
        grand_dr += dr; grand_cr += cr
        _f_row(ws, i, fill, [acct.account_code, acct.account_name, acct.account_type, dr, cr, dr - cr], num_cols=(4, 5, 6))

    total_r = len(accounts) + 6
    _f_row(ws, total_r, _F_DARK_FILL, ["", "TOTAL", "", grand_dr, grand_cr, grand_dr - grand_cr], num_cols=(4, 5, 6))

    for col, w in enumerate([14, 35, 12, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A6"


def _wpf_full_is_sheet(ws, trust, accounts, entries, label_from, label_to):
    from collections import defaultdict
    totals = defaultdict(lambda: {"dr": 0.0, "cr": 0.0})
    for e in entries:
        totals[e.account_code]["dr"] += e.debit
        totals[e.account_code]["cr"] += e.credit

    for r, val in [(1, trust.name), (2, "INCOME STATEMENT"), (3, f"Period: {label_from}  –  {label_to}")]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c = ws.cell(row=r, column=1, value=val)
        c.font = _F_WHITE_FONT if r < 3 else Font(italic=True, color="FFFFFF", size=9)
        c.fill, c.alignment = _F_DARK_FILL, _CENTER

    row = [5]

    def _section(title, accts, net_fn):
        ws.merge_cells(start_row=row[0], start_column=1, end_row=row[0], end_column=3)
        c = ws.cell(row=row[0], column=1, value=title)
        c.font, c.fill, c.alignment = _F_WHITE_FONT, _F_DARK_FILL, _LEFT
        row[0] += 1
        total = 0.0
        for i, acct in enumerate(accts):
            fill = _F_WHITE_FILL if i % 2 == 0 else _F_ALT_FILL
            net = net_fn(totals[acct.account_code]["dr"], totals[acct.account_code]["cr"])
            _f_row(ws, row[0], fill, [acct.account_code, acct.account_name, net], num_cols=(3,))
            total += net; row[0] += 1
        _f_row(ws, row[0], _F_DARK_FILL, ["", f"Total {title}", total], num_cols=(3,))
        row[0] += 2
        return total

    income_accts  = [a for a in accounts if a.account_type == "INCOME"]
    expense_accts = [a for a in accounts if a.account_type in ("EXPENSE",)]
    total_in  = _section("INCOME",   income_accts,  lambda dr, cr: cr - dr)
    total_exp = _section("EXPENSES", expense_accts, lambda dr, cr: dr - cr)

    _f_row(ws, row[0], _F_DARK_FILL, ["NET PROFIT / (LOSS)", "", total_in - total_exp], num_cols=(3,))

    for col, w in enumerate([14, 35, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A5"


def _wpf_full_bs_sheet(ws, trust, accounts, entries, label_from, label_to):
    from collections import defaultdict
    totals = defaultdict(lambda: {"dr": 0.0, "cr": 0.0})
    for e in entries:
        totals[e.account_code]["dr"] += e.debit
        totals[e.account_code]["cr"] += e.credit

    for r, val in [(1, trust.name), (2, "BALANCE SHEET"), (3, f"As at {label_to}")]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c = ws.cell(row=r, column=1, value=val)
        c.font = _F_WHITE_FONT if r < 3 else Font(italic=True, color="FFFFFF", size=9)
        c.fill, c.alignment = _F_DARK_FILL, _CENTER

    row = [5]

    def _section(title, accts, net_fn):
        ws.merge_cells(start_row=row[0], start_column=1, end_row=row[0], end_column=3)
        c = ws.cell(row=row[0], column=1, value=title)
        c.font, c.fill, c.alignment = _F_WHITE_FONT, _F_DARK_FILL, _LEFT
        row[0] += 1
        total = 0.0
        for i, acct in enumerate(accts):
            fill = _F_WHITE_FILL if i % 2 == 0 else _F_ALT_FILL
            net = net_fn(totals[acct.account_code]["dr"], totals[acct.account_code]["cr"])
            _f_row(ws, row[0], fill, [acct.account_code, acct.account_name, net], num_cols=(3,))
            total += net; row[0] += 1
        _f_row(ws, row[0], _F_DARK_FILL, ["", f"Total {title}", total], num_cols=(3,))
        row[0] += 2
        return total

    asset_accts  = [a for a in accounts if a.account_type == "ASSET"]
    liab_accts   = [a for a in accounts if a.account_type == "LIABILITY"]
    equity_accts = [a for a in accounts if a.account_type in ("EQUITY", "CAPITAL")]
    income_accts  = [a for a in accounts if a.account_type == "INCOME"]
    expense_accts = [a for a in accounts if a.account_type in ("EXPENSE",)]

    total_assets = _section("ASSETS",      asset_accts,  lambda dr, cr: dr - cr)
    total_liab   = _section("LIABILITIES", liab_accts,   lambda dr, cr: cr - dr)
    total_equity = _section("EQUITY / CAPITAL", equity_accts, lambda dr, cr: cr - dr)

    income_tot  = sum((totals[a.account_code]["cr"] - totals[a.account_code]["dr"]) for a in income_accts)
    expense_tot = sum((totals[a.account_code]["dr"] - totals[a.account_code]["cr"]) for a in expense_accts)
    net_profit  = income_tot - expense_tot

    _f_row(ws, row[0], _F_ALT_FILL, ["", "Net Profit / (Loss)", net_profit], num_cols=(3,))
    row[0] += 2
    _f_row(ws, row[0], _F_DARK_FILL, ["", "TOTAL LIABILITIES & EQUITY", total_liab + total_equity + net_profit], num_cols=(3,))
    row[0] += 1
    _f_row(ws, row[0], _F_DARK_FILL, ["", "TOTAL ASSETS", total_assets], num_cols=(3,))

    for col, w in enumerate([14, 35, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A5"


@router.get("/full")
def export_full(
    trust_id: int,
    year: Optional[int] = None,
    mode: str = "full",   # "full" | "ledger" | "tb"
    db: Session = Depends(get_db),
):
    """
    Export a workbook in exact WPF format (dark-green headers, one sheet per account).
    mode=full   → all account sheets + TB + IS + BS
    mode=ledger → account sheets only
    mode=tb     → Trial Balance sheet only
    """
    trust = db.query(Trust).filter(Trust.id == trust_id).first()
    if not trust:
        raise HTTPException(404, "Trust not found")

    from datetime import datetime as _dt
    if year:
        d_from_obj = _dt(year - 1, 7, 1).date()
        d_to_obj   = _dt(year, 6, 30).date()
        label_from = d_from_obj.isoformat()
        label_to   = d_to_obj.isoformat()
    else:
        d_from_obj = _dt(2000, 1, 1).date()
        d_to_obj   = date_type.today()
        label_from = "All time"
        label_to   = d_to_obj.isoformat()

    accounts = (
        db.query(AccountType)
        .filter(AccountType.trust_id == trust_id)
        .order_by(AccountType.account_type, AccountType.account_code)
        .all()
    )
    all_entries = (
        db.query(LedgerEntry)
        .filter(
            LedgerEntry.trust_id == trust_id,
            LedgerEntry.date >= d_from_obj,
            LedgerEntry.date <= d_to_obj,
        )
        .order_by(LedgerEntry.date, LedgerEntry.id)
        .all()
    )

    from collections import defaultdict
    by_acct: dict = defaultdict(list)
    for e in all_entries:
        by_acct[e.account_code].append(e)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if mode in ("full", "ledger"):
        for acct in accounts:
            safe = acct.account_code[:31].translate(str.maketrans("", "", r'/\*?[]'))
            ws = wb.create_sheet(safe or acct.account_code[:31])
            _wpf_full_account_sheet(ws, acct, by_acct.get(acct.account_code, []), trust.name)

    if mode in ("full", "tb"):
        ws_tb = wb.create_sheet("TB")
        _wpf_full_tb_sheet(ws_tb, trust, accounts, all_entries, label_from, label_to)

    if mode == "full":
        ws_is = wb.create_sheet("IS")
        _wpf_full_is_sheet(ws_is, trust, accounts, all_entries, label_from, label_to)
        ws_bs = wb.create_sheet("BS")
        _wpf_full_bs_sheet(ws_bs, trust, accounts, all_entries, label_from, label_to)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    yr_label  = str(year) if year else "all"
    mode_sfx  = "" if mode == "full" else f"-{mode}"
    filename  = f"{trust.code}-{yr_label}{mode_sfx}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
