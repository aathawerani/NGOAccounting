"""Backup system: DB snapshot + Excel exports for all trusts."""

import json
import os
import shutil
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
import openpyxl
from collections import defaultdict

from database import get_db
from models.models import AccountType, LedgerEntry, Trust
from routers.export_data import (
    _wpf_full_account_sheet,
    _wpf_full_tb_sheet,
    _wpf_full_is_sheet,
    _wpf_full_bs_sheet,
)

router = APIRouter(prefix="/api/backup", tags=["backup"])

# Backups folder sits next to the DB file (backend directory)
_BACKUP_DIR = Path(__file__).parent.parent / "Backups"
_META_FILE = _BACKUP_DIR / "last_backup.json"


def _ensure_backup_dir() -> None:
    _BACKUP_DIR.mkdir(parents=True, exist_ok=True)


def _load_meta() -> dict:
    if _META_FILE.exists():
        try:
            return json.loads(_META_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _save_meta(meta: dict) -> None:
    _META_FILE.write_text(json.dumps(meta, indent=2), encoding="utf-8")


def _build_trust_workbook(trust: Trust, db: Session) -> bytes:
    """Build a full WPF-format workbook for one trust and return raw bytes."""
    accounts = (
        db.query(AccountType)
        .filter(AccountType.trust_id == trust.id)
        .order_by(AccountType.account_type, AccountType.account_code)
        .all()
    )
    all_entries = (
        db.query(LedgerEntry)
        .filter(
            LedgerEntry.trust_id == trust.id,
            LedgerEntry.is_deleted == False,
        )
        .order_by(LedgerEntry.date, LedgerEntry.id)
        .all()
    )

    by_acct: dict = defaultdict(list)
    for e in all_entries:
        by_acct[e.account_code].append(e)

    from datetime import date as _date
    label_from = "All time"
    label_to = _date.today().isoformat()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for acct in accounts:
        safe = acct.account_code[:31].translate(str.maketrans("", "", r'/\*?[]'))
        ws = wb.create_sheet(safe or acct.account_code[:31])
        _wpf_full_account_sheet(ws, acct, by_acct.get(acct.account_code, []), trust.name)

    ws_tb = wb.create_sheet("TB")
    _wpf_full_tb_sheet(ws_tb, trust, accounts, all_entries, label_from, label_to)
    ws_is = wb.create_sheet("IS")
    _wpf_full_is_sheet(ws_is, trust, accounts, all_entries, label_from, label_to)
    ws_bs = wb.create_sheet("BS")
    _wpf_full_bs_sheet(ws_bs, trust, accounts, all_entries, label_from, label_to)

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get("/last")
def get_last_backup():
    """Return metadata about the most recent backup (or null if none)."""
    meta = _load_meta()
    return meta if meta else {"timestamp": None, "db_file": None, "excel_files": []}


@router.post("/create")
def create_backup(db: Session = Depends(get_db)):
    """
    Create a full backup:
    - DB snapshot (timestamped .db file)
    - Full Excel workbook for every trust
    """
    _ensure_backup_dir()
    ts = datetime.now()
    ts_str = ts.strftime("%Y%m%d_%H%M%S")
    ts_iso = ts.isoformat(timespec="seconds")

    # 1 — DB snapshot
    db_src = Path(__file__).parent.parent / "ngo_accounting.db"
    db_dest_name = f"ngo_accounting_{ts_str}.db"
    db_dest = _BACKUP_DIR / db_dest_name

    if db_src.exists():
        shutil.copy2(str(db_src), str(db_dest))
    else:
        db_dest_name = None  # no DB found

    # 2 — Excel workbooks per trust
    trusts = db.query(Trust).all()
    excel_files = []
    errors = []

    for trust in trusts:
        try:
            wb_bytes = _build_trust_workbook(trust, db)
            filename = f"{trust.code}_{ts_str}.xlsx"
            ((_BACKUP_DIR / filename)).write_bytes(wb_bytes)
            excel_files.append(filename)
        except Exception as exc:
            errors.append(f"{trust.code}: {exc}")

    # 3 — Persist metadata
    meta = {
        "timestamp": ts_iso,
        "db_file": db_dest_name,
        "excel_files": excel_files,
        "errors": errors,
        "backup_dir": str(_BACKUP_DIR),
    }
    _save_meta(meta)

    return {
        "ok": True,
        "timestamp": ts_iso,
        "db_file": db_dest_name,
        "excel_files": excel_files,
        "errors": errors,
        "backup_dir": str(_BACKUP_DIR),
    }
