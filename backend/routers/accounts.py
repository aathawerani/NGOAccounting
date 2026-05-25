import uuid
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy.orm import Session

from database import get_db
from models.models import AccountType, LedgerEntry, Trust

router = APIRouter(prefix="/api/accounts", tags=["accounts"])


# ── Schemas ───────────────────────────────────────────────────────────────────

class JournalEntryBody(BaseModel):
    trust_id: int
    date: date
    receipt_no: Optional[str] = None
    party_name: Optional[str] = None
    debit_account_code: str     # account that gets debited
    credit_account_code: str    # account that gets credited
    particulars: Optional[str] = None
    amount: float


# ── Helpers ───────────────────────────────────────────────────────────────────

def _serialize_entry(e: LedgerEntry) -> dict:
    return {
        "id": e.id,
        "trust_id": e.trust_id,
        "account_code": e.account_code,
        "date": e.date.isoformat() if e.date else None,
        "receipt_no": e.receipt_no,
        "party_name": e.party_name,
        "contra_account_code": e.contra_account_code,
        "particulars": e.particulars,
        "debit": e.debit,
        "credit": e.credit,
        "row_order": e.row_order,
        "account_key": e.account_key,
    }


def _serialize_account(a: AccountType) -> dict:
    return {
        "id": a.id,
        "trust_id": a.trust_id,
        "account_code": a.account_code,
        "account_name": a.account_name,
        "account_type": a.account_type,
        "is_certificate": a.is_certificate,
    }


# ── Routes ────────────────────────────────────────────────────────────────────

@router.get("/types")
def list_account_types(trust_id: Optional[int] = None, db: Session = Depends(get_db)):
    q = db.query(AccountType)
    if trust_id is not None:
        q = q.filter(AccountType.trust_id == trust_id)
    return [_serialize_account(a) for a in q.order_by(AccountType.account_name).all()]


@router.get("/ledger")
def get_ledger(
    trust_id: int,
    account_code: str,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
):
    """Return all ledger entries for one account, with opening balance and running balance."""
    acct = db.query(AccountType).filter(
        AccountType.trust_id == trust_id,
        AccountType.account_code == account_code,
    ).first()
    if not acct:
        raise HTTPException(status_code=404, detail="Account not found")

    # Opening balance: sum of all entries strictly before date_from
    opening_balance = 0.0
    if date_from:
        pre_entries = db.query(LedgerEntry).filter(
            LedgerEntry.trust_id == trust_id,
            LedgerEntry.account_code == account_code,
            LedgerEntry.is_deleted == False,
            LedgerEntry.date < date_from,
        ).all()
        opening_balance = round(sum(e.debit - e.credit for e in pre_entries), 2)

    q = db.query(LedgerEntry).filter(
        LedgerEntry.trust_id == trust_id,
        LedgerEntry.account_code == account_code,
        LedgerEntry.is_deleted == False,
    )
    if date_from:
        q = q.filter(LedgerEntry.date >= date_from)
    if date_to:
        q = q.filter(LedgerEntry.date <= date_to)

    entries = q.order_by(LedgerEntry.row_order, LedgerEntry.date, LedgerEntry.id).all()

    running = opening_balance
    result = []
    for e in entries:
        running += e.debit - e.credit
        row = _serialize_entry(e)
        row["balance"] = round(running, 2)
        result.append(row)

    return {
        "account": _serialize_account(acct),
        "opening_balance": opening_balance,
        "entries": result,
        "balance": round(running, 2),
    }


@router.get("/ledger/excel")
def get_ledger_excel(
    trust_id: int,
    account_code: str,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
):
    """Export account ledger as a styled xlsx file."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    acct = db.query(AccountType).filter(
        AccountType.trust_id == trust_id,
        AccountType.account_code == account_code,
    ).first()
    if not acct:
        raise HTTPException(status_code=404, detail="Account not found")

    data = get_ledger(trust_id, account_code, date_from, date_to, db)
    trust = db.query(Trust).filter(Trust.id == trust_id).first()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = acct.account_code[:31]

    dark_fill = PatternFill("solid", fgColor="1E293B")
    header_fill = PatternFill("solid", fgColor="334155")
    open_fill = PatternFill("solid", fgColor="EFF6FF")
    close_fill = PatternFill("solid", fgColor="F0FDF4")
    white_font = Font(color="FFFFFF", bold=True, size=10)
    bold = Font(bold=True, size=9)
    norm = Font(size=9)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(bottom=thin)

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = f"{trust.name if trust else ''} — {acct.account_name} ({acct.account_code})"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = center

    period = ""
    if date_from and date_to:
        period = f"{date_from.strftime('%d %b %Y')} to {date_to.strftime('%d %b %Y')}"
    elif date_from:
        period = f"From {date_from.strftime('%d %b %Y')}"
    elif date_to:
        period = f"Up to {date_to.strftime('%d %b %Y')}"
    else:
        period = "All Periods"

    ws.merge_cells("A2:F2")
    ws["A2"] = period
    ws["A2"].alignment = center

    headers = ["Date", "Entry Ref", "Particulars", "Debit (PKR)", "Credit (PKR)", "Balance (PKR)"]
    col_widths = [14, 18, 45, 16, 16, 16]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = white_font
        cell.fill = dark_fill
        cell.alignment = right if col >= 4 else center if col == 1 else Alignment()
        ws.column_dimensions[cell.column_letter].width = w

    row_num = 4
    # Opening balance row
    ob = data["opening_balance"]
    ws.cell(row=row_num, column=1, value="Opening Balance").font = bold
    ws.cell(row=row_num, column=1).fill = open_fill
    ws.cell(row=row_num, column=1).alignment = center
    for c in range(2, 6):
        ws.cell(row=row_num, column=c).fill = open_fill
    ws.cell(row=row_num, column=6, value=ob).fill = open_fill
    ws.cell(row=row_num, column=6).font = bold
    ws.cell(row=row_num, column=6).number_format = "#,##0"
    ws.cell(row=row_num, column=6).alignment = right
    ws.merge_cells(f"A{row_num}:C{row_num}")
    row_num += 1

    for entry in data["entries"]:
        ws.cell(row=row_num, column=1, value=entry["date"]).font = norm
        ws.cell(row=row_num, column=1).alignment = center
        ws.cell(row=row_num, column=2, value=entry.get("receipt_no") or entry.get("account_key") or "").font = norm
        ws.cell(row=row_num, column=3, value=entry.get("particulars") or "").font = norm
        for col, key in [(4, "debit"), (5, "credit"), (6, "balance")]:
            val = entry.get(key) or 0
            c = ws.cell(row=row_num, column=col, value=val if val else None)
            c.font = norm
            c.number_format = "#,##0"
            c.alignment = right
        ws.cell(row=row_num, column=6).fill = PatternFill("solid", fgColor="F8FAFC")
        row_num += 1

    # Closing balance row
    cb = data["balance"]
    ws.cell(row=row_num, column=1, value="Closing Balance").font = bold
    ws.cell(row=row_num, column=1).fill = close_fill
    ws.cell(row=row_num, column=1).alignment = center
    for c in range(2, 6):
        ws.cell(row=row_num, column=c).fill = close_fill
    ws.cell(row=row_num, column=6, value=cb).fill = close_fill
    ws.cell(row=row_num, column=6).font = bold
    ws.cell(row=row_num, column=6).number_format = "#,##0"
    ws.cell(row=row_num, column=6).alignment = right
    ws.merge_cells(f"A{row_num}:C{row_num}")

    ws.row_dimensions[3].height = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"ledger_{account_code}_{trust_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/journal", status_code=201)
def create_journal_entry(body: JournalEntryBody, db: Session = Depends(get_db)):
    """Create a dual-entry: debit_account DR, credit_account CR."""
    if not db.query(Trust).filter(Trust.id == body.trust_id).first():
        raise HTTPException(status_code=404, detail="Trust not found")

    for code in (body.debit_account_code, body.credit_account_code):
        if not db.query(AccountType).filter(
            AccountType.trust_id == body.trust_id,
            AccountType.account_code == code,
        ).first():
            raise HTTPException(status_code=404, detail=f"Account '{code}' not found")

    key = str(uuid.uuid4())[:8]

    debit_row = LedgerEntry(
        trust_id=body.trust_id,
        account_code=body.debit_account_code,
        date=body.date,
        receipt_no=body.receipt_no,
        party_name=body.party_name,
        contra_account_code=body.credit_account_code,
        particulars=body.particulars,
        debit=body.amount,
        credit=0.0,
        row_order=2,
        account_key=key,
    )
    credit_row = LedgerEntry(
        trust_id=body.trust_id,
        account_code=body.credit_account_code,
        date=body.date,
        receipt_no=body.receipt_no,
        party_name=body.party_name,
        contra_account_code=body.debit_account_code,
        particulars=body.particulars,
        debit=0.0,
        credit=body.amount,
        row_order=2,
        account_key=key,
    )
    db.add_all([debit_row, credit_row])
    db.commit()
    db.refresh(debit_row)
    db.refresh(credit_row)
    return [_serialize_entry(debit_row), _serialize_entry(credit_row)]


@router.delete("/journal/{account_key}")
def delete_journal_entry(account_key: str, trust_id: int, db: Session = Depends(get_db)):
    """Delete both legs of a dual entry by their shared account_key."""
    rows = db.query(LedgerEntry).filter(
        LedgerEntry.trust_id == trust_id,
        LedgerEntry.account_key == account_key,
    ).all()
    if not rows:
        raise HTTPException(status_code=404, detail="Entry not found")
    for r in rows:
        db.delete(r)
    db.commit()


@router.get("/transactions")
def list_transactions(
    trust_id: int,
    account_code: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    limit: int = 200,
    offset: int = 0,
    db: Session = Depends(get_db),
):
    """Return one row per transaction: debit side with its contra (credit) account.

    Imported entries have unique account_keys, so we use debit > 0 to pick the
    DR leg and contra_account_code for the CR side. Manually-created journal
    entries also follow this pattern.
    """
    q = db.query(LedgerEntry).filter(
        LedgerEntry.trust_id == trust_id,
        LedgerEntry.is_deleted == False,
        LedgerEntry.debit > 0,
    )
    if account_code:
        q = q.filter(
            (LedgerEntry.account_code == account_code) |
            (LedgerEntry.contra_account_code == account_code)
        )
    if date_from:
        q = q.filter(LedgerEntry.date >= date_from)
    if date_to:
        q = q.filter(LedgerEntry.date <= date_to)

    entries = q.order_by(LedgerEntry.date.desc(), LedgerEntry.id.desc()).all()

    txns = [
        {
            "account_key": e.account_key,
            "date": e.date.isoformat() if e.date else None,
            "receipt_no": e.receipt_no,
            "party_name": e.party_name,
            "particulars": e.particulars,
            "amount": round(e.debit, 2),
            "debit_account": e.account_code,
            "credit_account": e.contra_account_code or "—",
            "row_order": e.row_order,
        }
        for e in entries
    ]

    return {"total": len(txns), "transactions": txns[offset : offset + limit]}
