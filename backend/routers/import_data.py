"""
WPF Excel import — smart sync edition (v2).

Sheet layout (0-indexed Python rows / columns):
  row 0-1  : empty
  row 2    : "GENERAL LEDGER" title
  row 3    : empty
  row 4    : col[1]="NAME OF ACCOUNT :"  col[4]=account_name
  row 5    : col[1]="TYPE OF ACCOUNT:"   col[4]=account_type
  row 6    : col[1]="ACCOUNT CODE  :"    col[4]=account_code
  row 7-8  : empty / decorative
  row 9    : column headers — DATE | RECEIPT/VOUCHER NO | ACCOUNT CODE |
                              NAME OF TENANT | PARTICULARS | DEBIT | CREDIT | BALANCE
  row 10-11: empty
  row 12+  : transaction data (skip rows where col[0] DATE is empty)

Smart-sync rules:
  import_hash = MD5(f"{trust_id}|{account_code}|{particulars[:50]}")
  INSERT  : hash not found in DB
  UPDATE  : hash found, any of date/debit/credit/receipt_no/party_name/contra changed
  FLAG    : hash in DB but not in file  → is_deleted=True (never hard-deleted)
  RESTORE : hash was is_deleted=True, now re-appears → is_deleted=False
"""

import hashlib
import json
import uuid
from datetime import datetime, date as date_cls
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from sqlalchemy.orm import Session

from database import get_db
from models.models import AccountType, LedgerEntry, Tenant, Trust

router = APIRouter(prefix="/api/import", tags=["import"])

# ── Constants ─────────────────────────────────────────────────────────────────

_SKIP = {"TB", "IS", "BS", "DEP SCH", "DEP SCHEDULE", "DEP.SCH", "DEP. SCH",
         "DEPRECIATION SCHEDULE"}

_C_DATE        = 0
_C_VOUCHER     = 1
_C_CONTRA      = 2
_C_TENANT      = 3
_C_PARTICULARS = 4
_C_DEBIT       = 5
_C_CREDIT      = 6

_H_LABEL_COL = 1
_H_VALUE_COL = 4

_DEFAULT_DATA_ROW = 12

_SKIP_TENANT_VALUES = {"NAME OF TENANT", "TENANT", "PARTY", "NAME", ""}


# ── Low-level helpers ─────────────────────────────────────────────────────────

def _s(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _f(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, AttributeError):
        return 0.0


def _d(v, datemode: int = 0) -> Optional[date_cls]:
    if v is None:
        return None
    if isinstance(v, date_cls):
        return v
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, float) and v > 0:
        try:
            import xlrd
            return xlrd.xldate_as_datetime(v, datemode).date()
        except Exception:
            return None
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y",
                "%d-%b-%Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _make_hash(trust_id: int, account_code: str, particulars: Optional[str]) -> str:
    """Soft upsert key: same account + same description = same row."""
    raw = f"{trust_id}|{account_code}|{(particulars or '')[:50]}"
    return hashlib.md5(raw.encode("utf-8")).hexdigest()


# ── Workbook reader ───────────────────────────────────────────────────────────

class _Sheet:
    def __init__(self, name: str, rows: list[list], datemode: int = 0):
        self.name = name
        self.rows = rows
        self.datemode = datemode


def _load(content: bytes, filename: str) -> list[_Sheet]:
    if filename.lower().endswith(".xls"):
        try:
            import xlrd
        except ImportError:
            raise HTTPException(500, "xlrd not installed — run: pip install xlrd")
        try:
            wb = xlrd.open_workbook(file_contents=content)
            dm = wb.datemode
            sheets = []
            for sh in wb.sheets():
                rows = []
                for r in range(sh.nrows):
                    row = []
                    for c in range(sh.ncols):
                        cell = sh.cell(r, c)
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            try:
                                row.append(xlrd.xldate_as_datetime(cell.value, dm).date())
                            except Exception:
                                row.append(cell.value)
                        elif cell.ctype == xlrd.XL_CELL_EMPTY:
                            row.append(None)
                        else:
                            row.append(cell.value)
                    rows.append(row)
                sheets.append(_Sheet(sh.name, rows, dm))
            return sheets
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(400, f"Cannot open .xls file: {e}")
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
            sheets = [
                _Sheet(ws.title, [list(row) for row in ws.iter_rows(values_only=True)])
                for ws in wb.worksheets
            ]
            wb.close()
            return sheets
        except Exception as e:
            raise HTTPException(400, f"Cannot open .xlsx file: {e}")


# ── Header / transaction parsers ─────────────────────────────────────────────

def _parse_header(rows: list[list]) -> dict:
    meta = {"account_name": None, "account_type": None, "account_code": None}
    for row in rows[:13]:
        if not row or len(row) <= _H_VALUE_COL:
            continue
        label = str(row[_H_LABEL_COL] or "").strip().upper()
        value = _s(row[_H_VALUE_COL])
        if "NAME OF ACCOUNT" in label:
            meta["account_name"] = value
        elif "TYPE OF ACCOUNT" in label:
            meta["account_type"] = value.upper() if value else None
        elif "ACCOUNT CODE" in label:
            meta["account_code"] = value
    return meta


def _data_start(rows: list[list]) -> int:
    for i, row in enumerate(rows[:16]):
        if row and str(row[0] or "").strip().upper() == "DATE":
            return i + 3
    return _DEFAULT_DATA_ROW


def _parse_transactions(sheet: _Sheet, account_code: str) -> tuple[list[dict], list[str]]:
    rows = sheet.rows
    start = _data_start(rows)
    txns: list[dict] = []
    tenant_names: list[str] = []
    is_rt = _is_rent_water(account_code)

    for row in rows[start:]:
        if not row:
            continue
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        dt = _d(row[_C_DATE] if len(row) > _C_DATE else None, sheet.datemode)
        if dt is None:
            continue

        tenant = _s(row[_C_TENANT] if len(row) > _C_TENANT else None)
        if is_rt and tenant and tenant.upper() not in _SKIP_TENANT_VALUES:
            tenant_names.append(tenant)

        txns.append({
            "date":                dt,
            "receipt_no":          _s(row[_C_VOUCHER]     if len(row) > _C_VOUCHER     else None),
            "contra_account_code": _s(row[_C_CONTRA]      if len(row) > _C_CONTRA      else None),
            "party_name":          tenant,
            "particulars":         _s(row[_C_PARTICULARS] if len(row) > _C_PARTICULARS else None),
            "debit":               _f(row[_C_DEBIT]       if len(row) > _C_DEBIT       else 0),
            "credit":              _f(row[_C_CREDIT]      if len(row) > _C_CREDIT      else 0),
        })

    return txns, list(dict.fromkeys(tenant_names))


# ── Tenant helpers ────────────────────────────────────────────────────────────

def _is_rent_water(code: str) -> bool:
    if not code or len(code) < 2:
        return False
    prefix, second = code[0].upper(), code[1].upper()
    if prefix == "R" and second.isalnum() and "&" not in code:
        return True
    if prefix == "W" and second.isalnum() and code.upper() not in {"WHT", "WT"}:
        return True
    return False


def _plot_from_code(code: str) -> Optional[str]:
    return code[1:] if code and len(code) > 1 else None


# ── Trust detection ───────────────────────────────────────────────────────────

_TRUST_KEYWORDS: list[tuple[set, str]] = [
    ({"THAWER", "THARIA", "HTTT"}, "HTTT"),
    ({"HUSSAINI", "VAKIL", "HVHT"}, "HVHT"),
    ({"BAIT", "BURHANI", "BIB"},   "BIB"),
]


def _extract_trust_name(sheets: list[_Sheet]) -> Optional[str]:
    for sh in sheets:
        uname = sh.name.strip().upper()
        rows = sh.rows
        try:
            if uname == "TB" and len(rows) > 2:
                val = _s(rows[2][0] if rows[2] else None)
                if val:
                    return val
            elif uname == "IS" and len(rows) > 4:
                row4 = rows[4]
                val = _s(row4[1] if len(row4) > 1 else None)
                if val:
                    return val
            elif uname == "BS" and len(rows) > 0:
                row0 = rows[0]
                val = _s(row0[0] if row0 else None)
                if val:
                    return val
        except (IndexError, TypeError):
            continue
    return None


def _match_trust_keyword(detected: Optional[str], db: Session) -> Optional[dict]:
    if not detected:
        return None
    upper = detected.upper()
    for keywords, code in _TRUST_KEYWORDS:
        if any(kw in upper for kw in keywords):
            trust = db.query(Trust).filter(Trust.code == code).first()
            if trust:
                return {"id": trust.id, "code": trust.code, "name": trust.name}
    return None


# ── Date validation ───────────────────────────────────────────────────────────

def _validate_entries(trust_id: int, db: Session) -> int:
    """Run validation pass on all non-deleted entries; store JSON warnings. Returns count with warnings."""
    today = date_cls.today()

    # Build receipt_no → dates map for spread check
    from sqlalchemy import and_
    entries = (
        db.query(LedgerEntry)
        .filter(
            LedgerEntry.trust_id == trust_id,
            LedgerEntry.is_deleted == False,
        )
        .all()
    )

    receipt_dates: dict[str, list[date_cls]] = {}
    for e in entries:
        if e.receipt_no and e.date:
            receipt_dates.setdefault(e.receipt_no, []).append(e.date)

    warning_count = 0
    for e in entries:
        warns: list[str] = []
        if e.date:
            if e.date > today:
                warns.append("future_date")
            if e.date.year < 2000:
                warns.append("date_too_old")
            if e.receipt_no:
                dates = receipt_dates.get(e.receipt_no, [])
                if len(dates) > 1:
                    span = (max(dates) - min(dates)).days
                    if span > 30:
                        warns.append(f"voucher_spread_{span}d")

        new_val = json.dumps(warns) if warns else None
        if e.validation_warnings != new_val:
            e.validation_warnings = new_val
        if warns:
            warning_count += 1

    db.commit()
    return warning_count


# ── Tenant sync ───────────────────────────────────────────────────────────────

def _sync_tenants(trust_id: int, db: Session) -> tuple[int, int]:
    """Re-extract tenants from active rent/water contra entries; upsert by name."""
    entries = (
        db.query(LedgerEntry)
        .filter(
            LedgerEntry.trust_id == trust_id,
            LedgerEntry.party_name != None,
            LedgerEntry.party_name != "",
            LedgerEntry.is_deleted == False,
        )
        .all()
    )

    name_to_plot: dict[str, str] = {}
    for e in entries:
        if not _is_rent_water(e.contra_account_code or ""):
            continue
        tname = (e.party_name or "").strip()
        if not tname or tname.upper() in _SKIP_TENANT_VALUES:
            continue
        plot = _plot_from_code(e.contra_account_code) or ""
        name_to_plot[tname] = plot

    added = updated = 0
    for tname, plot in name_to_plot.items():
        existing = (
            db.query(Tenant)
            .filter(Tenant.trust_id == trust_id, Tenant.name == tname)
            .first()
        )
        if existing:
            if existing.plot_code != plot and plot:
                existing.plot_code = plot
                updated += 1
        else:
            db.add(Tenant(
                trust_id=trust_id,
                name=tname,
                plot_code=plot or None,
                space_type="SHOP",
                monthly_rent=0.0,
                water_charge=0.0,
                is_active=True,
            ))
            added += 1

    db.commit()
    return added, updated


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.post("/detect-trust")
async def detect_trust(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not file.filename.lower().endswith((".xls", ".xlsx")):
        raise HTTPException(400, "Only .xls and .xlsx files are supported")
    content = await file.read()
    sheets = _load(content, file.filename)
    detected_name = _extract_trust_name(sheets)
    matched = _match_trust_keyword(detected_name, db)
    all_trusts = db.query(Trust).order_by(Trust.code).all()
    return {
        "detected_name":      detected_name,
        "matched_trust_id":   matched["id"]   if matched else None,
        "matched_trust_code": matched["code"] if matched else None,
        "matched_trust_name": matched["name"] if matched else None,
        "confidence":         "high" if matched else "none",
        "all_trusts": [{"id": t.id, "code": t.code, "name": t.name} for t in all_trusts],
    }


@router.post("/preview")
async def preview(
    trust_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not file.filename.lower().endswith((".xls", ".xlsx")):
        raise HTTPException(400, "Only .xls and .xlsx files are supported")

    content = await file.read()
    sheets = _load(content, file.filename)

    # Load all existing hashes for this trust grouped by account_code
    existing_by_acct: dict[str, dict[str, list]] = {}
    for e in db.query(LedgerEntry).filter(
        LedgerEntry.trust_id == trust_id,
        LedgerEntry.import_hash != None,
        LedgerEntry.is_deleted == False,
    ).all():
        existing_by_acct.setdefault(e.account_code, {}).setdefault(e.import_hash, []).append(e)

    skipped: list[str] = []
    accounts: list[dict] = []
    all_tenants: list[dict] = []
    total_txns = 0
    total_est_inserts = total_est_updates = total_est_flags = 0

    for sh in sheets:
        if sh.name.strip().upper() in _SKIP:
            skipped.append(sh.name)
            continue

        meta = _parse_header(sh.rows)
        code = meta["account_code"] or sh.name.strip()
        txns, tenants = _parse_transactions(sh, code)
        total_txns += len(txns)

        if tenants:
            plot = _plot_from_code(code)
            for t in tenants:
                all_tenants.append({"name": t, "plot_code": plot})

        # Sync analysis for this account
        existing_hashes = existing_by_acct.get(code, {})
        seen: dict[str, int] = {}  # hash -> count in file
        for txn in txns:
            h = _make_hash(trust_id, code, txn["particulars"])
            seen[h] = seen.get(h, 0) + 1

        est_inserts = est_updates = est_flags = 0
        # Compare file hashes vs existing hashes
        consumed: dict[str, int] = {}
        for h, file_count in seen.items():
            db_entries = existing_hashes.get(h, [])
            for i in range(file_count):
                if i < len(db_entries):
                    est_updates += 1
                else:
                    est_inserts += 1
            consumed[h] = file_count

        for h, db_entries in existing_hashes.items():
            used = consumed.get(h, 0)
            leftover = len(db_entries) - used
            if leftover > 0:
                est_flags += leftover

        total_est_inserts += est_inserts
        total_est_updates += est_updates
        total_est_flags   += est_flags

        accounts.append({
            "sheet":             sh.name,
            "account_code":      code,
            "account_name":      meta["account_name"] or code,
            "account_type":      meta["account_type"] or "UNKNOWN",
            "transaction_count": len(txns),
            "est_inserts":       est_inserts,
            "est_updates":       est_updates,
            "est_flags":         est_flags,
        })

    seen_tenants: dict[str, dict] = {}
    for t in all_tenants:
        seen_tenants.setdefault(t["name"], t)

    return {
        "total_sheets":       len(sheets),
        "sheets_skipped":     skipped,
        "accounts_found":     len(accounts),
        "transactions_found": total_txns,
        "tenants_found":      len(seen_tenants),
        "sync_analysis": {
            "estimated_inserts": total_est_inserts,
            "estimated_updates": total_est_updates,
            "estimated_flags":   total_est_flags,
        },
        "accounts":  accounts,
        "tenants":   list(seen_tenants.values()),
        "warnings":  [],
    }


@router.post("/execute")
async def execute_import(
    trust_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not db.query(Trust).filter(Trust.id == trust_id).first():
        raise HTTPException(404, "Trust not found")
    if not file.filename.lower().endswith((".xls", ".xlsx")):
        raise HTTPException(400, "Only .xls and .xlsx files are supported")

    content = await file.read()
    sheets = _load(content, file.filename)

    accounts_created = accounts_updated = 0
    total_inserted = total_updated = total_flagged = total_restored = 0
    errors: list[str] = []
    log: list[dict] = []

    for sh in sheets:
        if sh.name.strip().upper() in _SKIP:
            log.append({"type": "skip", "sheet": sh.name})
            continue

        meta  = _parse_header(sh.rows)
        code  = meta["account_code"] or sh.name.strip()
        name  = meta["account_name"] or code
        atype = meta["account_type"] or "ASSET"

        # Upsert account type
        try:
            existing_acct = (
                db.query(AccountType)
                .filter(AccountType.trust_id == trust_id, AccountType.account_code == code)
                .first()
            )
            if existing_acct:
                existing_acct.account_name = name
                existing_acct.account_type = atype
                accounts_updated += 1
            else:
                db.add(AccountType(
                    trust_id=trust_id, account_code=code, account_name=name,
                    account_type=atype,
                    is_certificate=code in {"SSC", "DSC", "BEH", "BSC"},
                ))
                accounts_created += 1
            db.flush()
        except Exception as e:
            db.rollback()
            errors.append(f"Account {code}: {e}")
            continue

        # Build hash → [entries] lookup for this trust+account (import-originated only)
        existing_entries = (
            db.query(LedgerEntry)
            .filter(
                LedgerEntry.trust_id == trust_id,
                LedgerEntry.account_code == code,
                LedgerEntry.import_hash != None,  # only synced entries; manual ones stay untouched
            )
            .order_by(LedgerEntry.id)
            .all()
        )
        hash_to_entries: dict[str, list[LedgerEntry]] = {}
        for e in existing_entries:
            hash_to_entries.setdefault(e.import_hash, []).append(e)

        txns, tenant_names = _parse_transactions(sh, code)

        # Sync transactions
        consumed: dict[str, int] = {}  # hash -> how many from file matched
        inserted = updated = restored = 0

        for txn in txns:
            h = _make_hash(trust_id, code, txn["particulars"])
            db_list = hash_to_entries.get(h, [])
            idx = consumed.get(h, 0)

            if idx < len(db_list):
                # UPDATE existing entry
                e = db_list[idx]
                consumed[h] = idx + 1
                changed = False
                if e.date != txn["date"]:
                    e.date = txn["date"]; changed = True
                if abs((e.debit or 0) - txn["debit"]) > 0.001:
                    e.debit = txn["debit"]; changed = True
                if abs((e.credit or 0) - txn["credit"]) > 0.001:
                    e.credit = txn["credit"]; changed = True
                if e.receipt_no != txn["receipt_no"]:
                    e.receipt_no = txn["receipt_no"]; changed = True
                if e.party_name != txn["party_name"]:
                    e.party_name = txn["party_name"]; changed = True
                if e.contra_account_code != txn["contra_account_code"]:
                    e.contra_account_code = txn["contra_account_code"]; changed = True
                if e.particulars != txn["particulars"]:
                    e.particulars = txn["particulars"]; changed = True
                if e.is_deleted:
                    e.is_deleted = False
                    restored += 1
                    changed = True
                if changed:
                    updated += 1
            else:
                # INSERT new entry
                new_entry = LedgerEntry(
                    trust_id=trust_id,
                    account_code=code,
                    date=txn["date"],
                    receipt_no=txn["receipt_no"],
                    party_name=txn["party_name"],
                    contra_account_code=txn["contra_account_code"],
                    particulars=txn["particulars"],
                    debit=txn["debit"],
                    credit=txn["credit"],
                    row_order=2,
                    account_key=str(uuid.uuid4())[:8],
                    import_hash=h,
                    is_deleted=False,
                )
                db.add(new_entry)
                if h not in hash_to_entries:
                    hash_to_entries[h] = []
                hash_to_entries[h].append(new_entry)
                consumed[h] = consumed.get(h, 0) + 1
                inserted += 1

        # Flag entries not matched by this import (present in DB but absent from file)
        flagged = 0
        for h, db_list in hash_to_entries.items():
            used = consumed.get(h, 0)
            for e in db_list[used:]:
                if not e.is_deleted:
                    e.is_deleted = True
                    flagged += 1

        total_inserted += inserted
        total_updated  += updated
        total_flagged  += flagged
        total_restored += restored

        try:
            db.commit()
        except Exception as e:
            db.rollback()
            errors.append(f"Commit failed for sheet {sh.name}: {e}")
            continue

        log.append({
            "type":         "account",
            "sheet":        sh.name,
            "account_code": code,
            "account_name": name,
            "account_type": atype,
            "inserted":     inserted,
            "updated":      updated,
            "flagged":      flagged,
            "restored":     restored,
            # legacy field kept for backward compat
            "transactions": inserted + updated,
            "tenants":      len(tenant_names),
        })

    # Tenant sync
    tenants_added, tenants_updated = _sync_tenants(trust_id, db)

    # Validation pass
    validation_warnings = _validate_entries(trust_id, db)

    return {
        "transactions_inserted":  total_inserted,
        "transactions_updated":   total_updated,
        "transactions_flagged":   total_flagged,
        "transactions_restored":  total_restored,
        "accounts_created":       accounts_created,
        "accounts_updated":       accounts_updated,
        "tenants_added":          tenants_added,
        "tenants_updated":        tenants_updated,
        "validation_warnings":    validation_warnings,
        "errors":                 errors,
        "log":                    log,
    }


@router.get("/warnings")
def get_warnings(trust_id: int, db: Session = Depends(get_db)):
    """Return all non-deleted entries with validation warnings."""
    entries = (
        db.query(LedgerEntry)
        .filter(
            LedgerEntry.trust_id == trust_id,
            LedgerEntry.is_deleted == False,
            LedgerEntry.validation_warnings != None,
        )
        .order_by(LedgerEntry.account_code, LedgerEntry.date, LedgerEntry.id)
        .all()
    )
    return [
        {
            "id":            e.id,
            "account_code":  e.account_code,
            "date":          e.date.isoformat() if e.date else None,
            "receipt_no":    e.receipt_no,
            "party_name":    e.party_name,
            "particulars":   e.particulars,
            "debit":         e.debit,
            "credit":        e.credit,
            "warnings":      json.loads(e.validation_warnings or "[]"),
        }
        for e in entries
        if e.validation_warnings and e.validation_warnings not in ("null", "[]")
    ]


@router.get("/flagged")
def get_flagged(trust_id: int, db: Session = Depends(get_db)):
    """Return entries flagged as deleted (present in DB but absent from latest import)."""
    entries = (
        db.query(LedgerEntry)
        .filter(LedgerEntry.trust_id == trust_id, LedgerEntry.is_deleted == True)
        .order_by(LedgerEntry.account_code, LedgerEntry.date, LedgerEntry.id)
        .all()
    )
    return [
        {
            "id":                  e.id,
            "account_code":        e.account_code,
            "date":                e.date.isoformat() if e.date else None,
            "receipt_no":          e.receipt_no,
            "party_name":          e.party_name,
            "particulars":         e.particulars,
            "debit":               e.debit,
            "credit":              e.credit,
            "contra_account_code": e.contra_account_code,
        }
        for e in entries
    ]
